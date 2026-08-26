#!/usr/bin/env python3
"""PT. Surya Graha Pasaraya price lists -> Akay Offers import CSV.

The generic intake_pricelist.py assumes one header row and one product per
row. The Surya Graha lists do not fit that: the L'Oreal/Garnier sheet carries
brand-section banner rows inside the data, and the Ellips sheet uses merged
cells so a product name (and sometimes its variant list) covers several
packaging rows. This adapter handles both layouts, applies the trading
margin and the offer expiry, and writes the import CSV plus a skipped-rows
file - no row is ever dropped silently.

Usage:
  python3 intake_surya_graha.py --loreal <xlsx> --ellips <xlsx> --out <dir> \
      [--margin 5] [--expiry-days 30] [--currency USD] [--today YYYY-MM-DD]
"""
import argparse, csv, os, re, datetime

SUPPLIER = "PT. Surya Graha Pasaraya"
TERMS = "FOB Jakarta"
ORIGIN = "Indonesia"
CATEGORY = "Toiletries"          # the catalogue has no Cosmetics category

COLUMNS = [
    "Source List", "Source Row", "Supplier Name", "Supplier Ref", "Barcode",
    "Brand", "Range", "Product Name", "Public Product Description", "Variant",
    "Category", "Supplier Category", "Public Spec", "Pack", "Unit Size",
    "Volume ML", "PCS/Case", "Unit Type",
    "Buy Price", "Buy Price Unit", "Currency", "Price Type",
    "Margin %", "Sell Price Per Case", "Sell Price Per Unit",
    "Price Display", "Price Per Unit & Case",
    "Stock Display", "Stock Cases", "MOQ", "Incoterm", "Public Terms",
    "Origin Country", "Warehouse", "BBD", "Offer Expiry", "Public Listing",
    "Featured", "Supplier Remarks", "Review", "Notes",
]

# --- name cleaning -------------------------------------------------------
# Only codes whose meaning is unambiguous are expanded. Anything left over is
# reported through the Review column instead of being guessed at.
BRAND_BY_SIGN = {"GAR": "Garnier", "OAP": "L'Oreal Paris",
                 "MYB": "Maybelline", "MRE": "L'Oreal Paris"}
LEAD_CODES = {"GAR": "Garnier", "GA": "Garnier", "GRN": "Garnier",
              "LOR": "L'Oreal", "OAP": "L'Oreal", "MYB": "Maybelline"}
# Division/market/pack codes the supplier carries in its SKU text. They say
# nothing about the product, so they are dropped wherever they appear.
DROP_CODES = {"DEX", "MU", "GA", "GRN", "AS", "EB", "EBT", "ET", "ETB", "ETC", "CND", "PCR",
              "V1", "V2", "OS", "NEW", "LOR", "OAP", "GAR", "MYB"}
TRAILING_CODES = DROP_CODES | {"X"}
# Only codes whose L'Oreal/Garnier/Maybelline meaning is unambiguous from the
# range column are expanded. Everything else is reported, never guessed.
WORDS = {
    "SHP": "Shampoo", "CDR": "Conditioner", "COND": "Conditioner",
    "CONDI": "Conditioner", "CRM": "Cream", "CRE": "Cream", "CREME": "Creme",
    "SRM": "Serum", "SER": "Serum", "ESS": "Essence", "FDT": "Foundation",
    "PWD": "Powder", "PWR": "Power", "LIPSTK": "Lipstick", "MCLR": "Micellar",
    "WTR": "Water", "CLSN": "Cleansing", "SCH": "Sachet", "SCHT": "Sachet",
    "COMP": "Complete", "COM": "Compact", "BRIT": "Bright", "BRT": "Bright",
    "BR": "Bright", "WHT": "White", "EXFO": "Exfoliating", "REF": "Refill",
    "ULT": "Ultra", "MAT": "Matte", "CUSH": "Cushion", "TB": "Tube",
    "DEFD": "Defender", "LTD": "Limited Edition", "BLST": "Blister",
    "CLR": "Color", "FRE": "Fresh", "FRH": "Fresh", "FRSH": "Fresh",
    "MSCR": "Mascara", "SACH": "Sachet", "AMP": "Ampoule", "VIT": "Vitamin",
    "HYD": "Hydra", "MIN": "Mineral", "NAT": "Natural", "FASH": "Fashion",
    "SMOOT": "Smooth", "INTES": "Intense", "INTEN": "Intense",
    "PROTCT": "Protect", "CORCT": "Correct",
    # ranges
    "ELS": "Elseve", "MEX": "Men Expert", "EXC": "Excellence",
    "SL": "Studio Line", "REV": "Revitalift", "WP": "White Perfect",
    "UVP": "UV Perfect", "UVD": "UV Defender", "WA": "White Activ",
    "TR5": "Total Repair 5", "FR": "Fall Resist", "SI": "Smooth Intense",
    "SM.INT": "Smooth Intense", "EXO": "Extraordinary Oil",
    "CP": "Color Protect", "COL": "Color", "HA": "Hyaluronic",
    "RENUT": "Re-Nutrition", "CR": "Color Riche", "GLYCO": "Glycolic",
    "BC": "Bright Complete", "SG": "Sakura Glow", "GSN": "Skin Naturals",
    "GMN": "Men", "GCN": "Color Naturals", "TLOC": "Turbo Light Oil Control",
    "TL": "Turbo Light", "INF": "Infallible", "INFA": "Infallible",
    "INFALB": "Infallible", "INFAILLIB": "Infallible", "HYD.FRESH": "Hydrafresh",
    "GWP": "Gift Pack", "BDL": "Bundle", "VS": "Value Set", "PCS": "pcs",
    "PC": "pc",
}
# Tokens that are ordinary words already - never counted as "unexplained".
KNOWN_WORDS = set("""ACID ACNE ACNO ACTIVE AMPOULE ANTI AQUA BABY BAG BALM BATH BB BEAUTY BIG BLACK
BLONDE BLOSSOM BLUE BODY BOGOF BOOSTER BOX BREEZE BRIGHT BRIGHTENING BROW BROWN BRUSH CARE
CHARCOAL CLAY CLEAN CLEANSER CLEANSING COLOR COMB COMPLETE CONDITIONER COOL COPPER CREAM CURL
DARK DAY DEEP DEO DEODORANT DERM DOUBLE DRY DUO ESSENCE EYE FACE FASHION FIGHT FIT FLEX FOAM
FOR FREE FRESH GEL GLOSS GLOW GOLD GREEN HAIR HAND ICY INK JAR KIT LE LIGHT LIP LIPSTICK LOTION
LUMI MASK MASCARA MATTE ME MEN MILK MIST MOISTURE MONO MOUSSE NATURAL NIGHT NUDE OIL PACK PEEL
PEELING PINK PORE POWDER PROTECT PURE RED REPAIR ROLL ROSE SCRUB SERUM SET SHAMPOO SHINE SILKY
SKIN SLIM SMOOTH SOFT SPOT SPRAY SUN TINT TONER TREATMENT TUBE UV VITAMIN WATER WAX WHIP WHITE
WITH AND YUZU ASH PLUM TEA BOMB UP ON IN KITS KIT PUMP LOW BIG BOOST SHAVE SHAVER STAR
CARAMEL CHERRY BERRY PEACH LATTE ROSE GOLD SAND IVORY NUDE PORCELAIN""".split())

# A handful of codes mean different things by department; resolve them from the
# supplier's own range/category columns rather than guessing one meaning.
BY_CONTEXT = {"CR": (("MAKEUP",), "Color Riche", "Cream")}

VOL_RE = re.compile(r"(\d+(?:[.,]\d+)?)\s*(ML|MLT|L|GR|G)\b", re.I)
MULTI_RE = re.compile(r"(\d+)\s*[xX]\s*(\d+(?:[.,]\d+)?)\s*(ML|MLT|GR|G)\b", re.I)
PLUS_RE = re.compile(r"(\d+(?:[.,]\d+)?)\s*\+\s*(\d+(?:[.,]\d+)?)\s*(ML|MLT|GR|G)\b", re.I)


def unit_size(name, first=False):
    """-> (display size, volume in ml or None, note). Weights stay out of Volume ML.
    `first` takes the leading size instead of the trailing one - a bundle name
    ends with the size of the free item, not of the product being priced."""
    s = str(name)
    m = PLUS_RE.search(s)          # "300+100ml" - main pack plus a free extra
    if m:
        a, b = float(m.group(1).replace(",", ".")), float(m.group(2).replace(",", "."))
        u = m.group(3).lower()
        txt = f"{m.group(1)}+{m.group(2)}{'ml' if u.startswith('ml') else 'g'}"
        return txt, (a + b if u.startswith("ml") else None), "bonus-pack size (main + extra)"
    m = MULTI_RE.search(s)         # "2X15ML" - two units inside one pack
    if m:
        cnt, size = int(m.group(1)), float(m.group(2).replace(",", "."))
        u = m.group(3).lower()
        txt = f"{cnt} x {m.group(2)}{'ml' if u.startswith('ml') else 'g'}"
        return txt, (cnt * size if u.startswith("ml") else None), "multi-unit pack"
    m = None
    for m in VOL_RE.finditer(s):
        if first:
            break                   # bundles: the first size is the main item
        pass                        # otherwise the last volume is the pack size
    if m:
        val = float(m.group(1).replace(",", "."))
        u = m.group(2).lower()
        if u in ("ml", "mlt"):
            return f"{val:g}ml", val, ""
        if u == "l":
            return f"{val:g}L", val * 1000, ""
        return f"{val:g}g", None, ""
    return "", None, ""


def clean_name(raw, brand, context=""):
    """Expand the supplier's SKU shorthand as far as it is safe to, and report
    whatever is left so a human can finish the name before publishing."""
    s = re.sub(r"\s+", " ", str(raw)).strip(" -")
    s = re.sub(r"(\d)\.0\s*(ML|MLT|GR)\b", r"\1\2", s, flags=re.I)   # 100.0MLT -> 100MLT
    s = re.sub(r"(?i)(\d)\s*MLT\b", r"\1ML", s)                       # 100MLT  -> 100ML
    tokens = [t for t in s.split() if t.upper().strip(".,-") not in DROP_CODES]
    while tokens and tokens[-1].upper().strip("-") in TRAILING_CODES:
        tokens.pop()
    out, unexplained = [], []
    for t in tokens:
        key = t.upper().strip(".,-")
        if key in BY_CONTEXT:
            where, hit, miss = BY_CONTEXT[key]
            out.append(hit if any(w in context.upper() for w in where) else miss)
        elif key in WORDS:
            out.append(WORDS[key])
        elif key in KNOWN_WORDS:
            out.append(t.title())
        elif re.fullmatch(r"[\d.,+#x/&%-]*(ML|G|GR|L|PCS?|H|IN|CM)?", key, re.I):
            out.append(t.lower() if re.search(r"\d", key) else t)
        elif re.fullmatch(r"SPF\s*\d*", key):
            out.append(key)
        elif len(key) <= 4 and key.isalpha():
            unexplained.append(key)          # short code we will not guess at
            out.append(key)
        else:
            out.append(t.title() if t.isupper() else t)
    name = re.sub(r"\s+", " ", " ".join(out)).strip(" -")
    name = re.sub(r"(?i)\b(\d+(?:\.\d+)?)\s*ml\b", r"\1ml", name)
    name = re.sub(r"(?i)\bspf\s*(\d+)", r"SPF\1", name)
    return name, sorted(set(unexplained))


def price_block(buy_case, pack, margin, currency):
    """Buy price -> sell price, per case and per unit, plus the display strings.
    Unit figures keep enough decimals that case == unit x pack stays inside the
    validator's 2% tolerance for cheap sachet packs."""
    sell_case = round(buy_case * (1 + margin / 100), 2)
    raw_unit = sell_case / pack if pack else None
    if raw_unit is None:
        return sell_case, None, "", ""
    dp = 2 if raw_unit >= 0.5 else 3 if raw_unit >= 0.05 else 4
    sell_unit = round(raw_unit, dp)
    disp_case = f"{currency} {sell_case:.2f}"
    detail = f"{disp_case}/case ({pack:g}pk) · {currency} {sell_unit:.{dp}f}/unit"
    return sell_case, sell_unit, disp_case, detail


def base_row(**kw):
    row = {c: "" for c in COLUMNS}
    row.update(kw)
    return row


def finish(row, buy_case, pack, args, unexplained, notes, review):
    sell_case, sell_unit, disp, detail = price_block(buy_case, pack, args.margin, args.currency)
    buy_unit = round(buy_case / pack, 4) if pack else ""
    if unexplained:
        review.append("name codes: " + ", ".join(unexplained))
    row.update({
        "Supplier Name": SUPPLIER, "Category": CATEGORY,
        "Buy Price": round(buy_case, 2), "Buy Price Unit": buy_unit,
        "Currency": args.currency, "Price Type": "Per Case",
        "Margin %": args.margin, "Sell Price Per Case": sell_case,
        "Sell Price Per Unit": sell_unit, "Price Display": disp,
        "Price Per Unit & Case": detail,
        "Stock Display": "Enquire", "Stock Cases": "",
        "Incoterm": TERMS, "Public Terms": TERMS, "Origin Country": ORIGIN,
        "Warehouse": "Jakarta", "Offer Expiry": args.expiry,
        "Public Listing": "No", "Featured": "No",
        "Review": "; ".join(review), "Notes": "; ".join(n for n in notes if n),
    })
    return row


# --- L'Oreal / Garnier / Maybelline sheet --------------------------------
def parse_loreal(path, args, out, skipped):
    from openpyxl import load_workbook
    ws = load_workbook(path, data_only=True).active
    for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        sign, rng, kat, barcode, code, name, pack, unit, price, remark = \
            [r[j] if j < len(r) else None for j in range(10)]
        cells = [c for c in (sign, rng, kat, barcode, code, name, pack, unit, price)
                 if c not in (None, "", "-")]
        label = re.sub(r"\s+", " ", str(name or code or "")).strip()
        if not cells:
            if any(c not in (None, "") for c in (sign, rng, kat, barcode, code, name)):
                skipped.append({"Source Row": i, "Source List": "LOREAL_GARNIER",
                                "Ref": "", "Name": label,
                                "Reason": "filler row (dashes only)"})
            continue
        if not isinstance(price, (int, float)) or price == 0:
            # brand banner rows carry a name and nothing else; real products
            # with a zero price are a supplier gap and must be reported.
            has_sku = str(barcode or "").strip() not in ("", "-")
            kind = ("no FOB price on the list" if has_sku else
                    "section banner" if label else "blank spacer row")
            ref = str(code or "").strip()
            skipped.append({"Source Row": i, "Source List": "LOREAL_GARNIER",
                            "Ref": "" if ref == "-" else ref,
                            "Name": label, "Reason": kind})
            continue
        if not isinstance(pack, (int, float)) or pack <= 0:
            skipped.append({"Source Row": i, "Source List": "LOREAL_GARNIER",
                            "Ref": str(code or ""), "Name": label,
                            "Reason": "no units-per-case"})
            continue
        brand = BRAND_BY_SIGN.get(str(sign or "").strip().upper(), "")
        if not brand:
            brand = "L'Oreal Paris" if "OREAL" in str(rng or "").upper() else str(rng or "").title()
        public, unexplained = clean_name(name, brand, f"{rng or ''} {kat or ''}")
        bundle = bool(re.search(r"(?i)\b(VS|BDL|BOGOF|FREE|GWP)\b|\+", str(name)))
        size_txt, ml, size_note = unit_size(name, first=bundle)
        pack = int(pack)
        unit_type = {"PCS": "Piece", "BOKS": "Box", "BOTTLE": "Bottle", "TUBE": "Tube",
                     "SACHET": "Sachet", "JAR": "Jar", "BOX": "Box", "PACK": "Pack",
                     "PUMP": "Pump", "AMPOULE": "Ampoule", "PEN": "Pen", "CAN": "Can",
                     "BUNDLE": "Bundle", "BAG": "Bag"}.get(str(unit or "").strip().upper(), "Piece")
        spec = f"{pack} x {size_txt}" if size_txt else str(pack)
        pack_spec = f"{pack} x {size_txt}" if size_txt else f"{pack} x {unit_type.lower()}"
        review = [] if size_txt else ["no unit size on the supplier line"]
        notes = [size_note]
        if bundle:
            review.append("bundle/value set - confirm what the case contains")
        rem = str(remark or "").strip()
        if re.search(r"sell through", rem, re.I):
            review.append("supplier marks it sell-through only")
        row = base_row(**{
            "Source List": "LOREAL_GARNIER", "Source Row": i,
            "Supplier Ref": str(code or "").strip(),
            "Barcode": str(int(barcode)) if isinstance(barcode, (int, float)) else str(barcode or "").strip(),
            "Brand": brand, "Range": str(rng or "").title(),
            "Product Name": re.sub(r"\s+", " ", str(name)).strip(),
            "Public Product Description": f"{brand} {public}".strip(),
            "Supplier Category": str(kat or "").title(),
            "Public Spec": spec, "Pack": pack_spec, "Unit Size": size_txt,
            "Volume ML": ml if ml else "", "PCS/Case": pack, "Unit Type": unit_type,
            "Supplier Remarks": rem,
        })
        out.append(finish(row, float(price), pack, args, unexplained, notes, review))


# --- Ellips sheet --------------------------------------------------------
ID_WORDS = {"Botol": "Bottle", "Renceng": "Strip",
            "Perawatan Rambut Rusak": "Damaged Hair Treatment",
            "Perawatan Rambut Hitam": "Black Hair Treatment"}


def parse_ellips(path, args, out, skipped):
    from openpyxl import load_workbook
    ws = load_workbook(path, data_only=True).active
    # merged cells hold the real value only in the top-left cell; spread it so a
    # continuation row inherits exactly what the supplier meant it to inherit.
    filled = {}
    for rng in ws.merged_cells.ranges:
        val = ws.cell(rng.min_row, rng.min_col).value
        for rr in range(rng.min_row, rng.max_row + 1):
            for cc in range(rng.min_col, rng.max_col + 1):
                filled[(rr, cc)] = val

    def cell(rr, cc):
        return filled.get((rr, cc), ws.cell(rr, cc).value)

    for i in range(14, ws.max_row + 1):
        name, variant, vol, pack, packtxt, packaging, price = \
            (cell(i, c) for c in (3, 4, 5, 6, 7, 8, 9))
        if not any(v not in (None, "") for v in (name, variant, vol, price)):
            continue
        label = re.sub(r"\s+", " ", str(name or "")).strip()
        if not isinstance(price, (int, float)) or price == 0:
            skipped.append({"Source Row": i, "Source List": "ELLIPS", "Ref": "",
                            "Name": label, "Reason": "no FOB price on the list"})
            continue
        if not isinstance(pack, (int, float)) or pack <= 0:
            skipped.append({"Source Row": i, "Source List": "ELLIPS", "Ref": "",
                            "Name": label, "Reason": "no units-per-case"})
            continue
        pack = int(pack)
        notes, review = [], []
        variant = re.sub(r"\s*,\s*", ", ", re.sub(r"\s+", " ", str(variant or "")).strip())
        for src, dst in ID_WORDS.items():
            if src in variant:
                variant = variant.replace(src, dst)
                notes.append(f"variant translated from Indonesian ({src})")
        vol_txt = re.sub(r"\s+", " ", str(vol or "")).strip()
        size_txt, ml, size_note = unit_size(vol_txt)
        notes.append(size_note)
        if not size_txt:                      # "Jar", "Blister", "Sachet", "Hanger"
            notes.append(f"supplier gives format only: {vol_txt}")
            review.append("no unit size on the supplier line")
        unit_type = ID_WORDS.get(str(packaging or "").strip(),
                                 str(packaging or "").strip().title()) or "Piece"
        packtxt = re.sub(r"\s+", " ", str(packtxt or "")).strip()
        if "Renceng" in packtxt:
            notes.append(f"supplier pack text: {packtxt} (12 strips = 144 sachets)")
        title = re.sub(r"\s+", " ", str(name or "")).strip().title().replace("Ellips", "Ellips")
        public = title
        if vol_txt and not size_txt:
            public = f"{title} ({vol_txt})"     # Jar / Blister / Sachet formats
        elif size_txt:
            public = f"{title} {size_txt}"
        spec = f"{pack} x {size_txt}" if size_txt else f"{pack} x {vol_txt or unit_type}"
        row = base_row(**{
            "Source List": "ELLIPS", "Source Row": i, "Supplier Ref": "", "Barcode": "",
            "Brand": "Ellips", "Range": "Ellips Hair Care",
            "Product Name": label, "Public Product Description": public,
            "Variant": variant, "Supplier Category": "Hair Care",
            "Public Spec": spec, "Pack": spec, "Unit Size": size_txt,
            "Volume ML": ml if ml else "", "PCS/Case": pack, "Unit Type": unit_type,
            "Supplier Remarks": packtxt,
        })
        out.append(finish(row, float(price), pack, args, [], notes, review))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--loreal"); ap.add_argument("--ellips")
    ap.add_argument("--out", default=".")
    ap.add_argument("--margin", type=float, default=5.0)
    ap.add_argument("--expiry-days", type=int, default=30)
    ap.add_argument("--currency", default="USD")
    ap.add_argument("--today", default=datetime.date.today().isoformat())
    args = ap.parse_args()
    args.expiry = (datetime.date.fromisoformat(args.today)
                   + datetime.timedelta(days=args.expiry_days)).isoformat()

    out, skipped = [], []
    if args.loreal:
        parse_loreal(args.loreal, args, out, skipped)
    if args.ellips:
        parse_ellips(args.ellips, args, out, skipped)

    # the supplier lists some SKUs twice (different distributor codes, same
    # product and price) - flag them so the catalogue does not get twin cards.
    seen = {}
    for r in out:
        seen.setdefault((r["Brand"], r["Public Product Description"],
                         r["PCS/Case"], r["Sell Price Per Case"]), []).append(r)
    dupes = 0
    for group in seen.values():
        if len(group) > 1:
            dupes += len(group)
            for r in group:
                flag = f"duplicate line in the supplier list ({len(group)}x)"
                r["Review"] = f"{r['Review']}; {flag}" if r["Review"] else flag

    os.makedirs(args.out, exist_ok=True)
    imp = os.path.join(args.out, "surya-graha-akay-import.csv")
    skp = os.path.join(args.out, "surya-graha-skipped.csv")
    with open(imp, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=COLUMNS); w.writeheader(); w.writerows(out)
    with open(skp, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["Source List", "Source Row", "Ref", "Name", "Reason"])
        w.writeheader(); w.writerows(skipped)

    print(f"OK: {len(out)} offer rows -> {imp}")
    print(f"    margin {args.margin:g}% | expiry {args.expiry} ({args.expiry_days} days from {args.today}) | {args.currency}")
    flagged = [o for o in out if o["Review"]]
    print(f"REVIEW: {len(flagged)} rows carry a review flag")
    reasons = {}
    for o in flagged:
        for part in o["Review"].split("; "):
            key = part.split(":")[0]
            reasons[key] = reasons.get(key, 0) + 1
    for k, v in sorted(reasons.items(), key=lambda kv: -kv[1]):
        print(f"    {v:5d}  {k}")
    print(f"    of which {dupes} rows are duplicate supplier lines")
    print(f"SKIPPED: {len(skipped)} rows -> {skp}")
    why = {}
    for s in skipped:
        why[s["Reason"]] = why.get(s["Reason"], 0) + 1
    for k, v in sorted(why.items(), key=lambda kv: -kv[1]):
        print(f"    {v:5d}  {k}")


if __name__ == "__main__":
    main()
